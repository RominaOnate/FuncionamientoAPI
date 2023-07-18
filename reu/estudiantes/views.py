from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from django.http import HttpResponse
from django.forms import modelform_factory
from rest_framework import viewsets, permissions, generics

from estudiantes.models import Estudiante
from estudiantes.forms import EstudianteFormulario
from openpyxl import Workbook
from estudiantes.serializers import EstudianteSerializer


# Create your views here.
def agregar_estudiante(request):
    pagina = loader.get_template('agregar_estudiantes.html')
    if request.method == 'GET':
        formulario = EstudianteFormulario
    elif request.method == 'POST':
        formulario = EstudianteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_estudiante(request, idEstudiante):
    pagina = loader.get_template('ver_estudiante.html')
    #estudiante = Estudiante.objects.get(pk=idEstudiante)
    estudiante = get_object_or_404(Estudiante, pk=idEstudiante)
    mensaje = {'estudiante': estudiante}
    return HttpResponse(pagina.render(mensaje, request))

def editar_estudiante(request, idEstudiante):
    pagina = loader.get_template('editar_estudiante.html')
    estudiante = get_object_or_404(Estudiante, pk=idEstudiante)
    if request.method == "GET":
        formulario = EstudianteFormulario(instance=estudiante)
    elif request.method == "POST":
        formulario = EstudianteFormulario(request.POST, instance=estudiante)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')

    mensaje = {'formulario': formulario}

    return HttpResponse(pagina.render(mensaje, request))

def eliminar_estudiante(request, idEstudiante):
    estudiante = get_object_or_404(Estudiante, pk=idEstudiante)
    if estudiante:
        estudiante.delete()
        return redirect('inicio')

def generar_reporte(request):
    # Obtenemos todos los estudiantes de nuestra base de datos
    estudiantes = Estudiante.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE ESTUDIANTES'
    ws['B1'] = 'REPORTE DE ESTUDIANTES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'CEDULA'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'NOMBRE'
    ws['E3'] = 'SEXO '
    ws['F3'] = 'EMAIL '
    ws['G3'] = 'FACULTAD '
    ws['H3'] = 'CARRERA '
    ws['I3'] = 'VIGENCIA '
    cont = 4
    # Recorremos el conjunto de estudinates y vamos escribiendo cada uno de los datos en las celdas
    for estudiante in estudiantes:
        ws.cell(row=cont, column=2).value = estudiante.cedula
        ws.cell(row=cont, column=3).value = estudiante.apellido
        ws.cell(row=cont, column=4).value = estudiante.nombre
        ws.cell(row=cont, column=5).value = estudiante.sexo
        ws.cell(row=cont, column=6).value = estudiante.email
        ws.cell(row=cont, column=7).value = estudiante.facultad.nombre_facultad if estudiante.facultad else ""
        ws.cell(row=cont, column=8).value = estudiante.carrera.nombre_carrera
        ws.cell(row=cont, column=9).value = estudiante.vigencia

        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteEstudiantesExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

class EstudianteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer
    permission_classes = [permissions.IsAuthenticated]

class EstudianteDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer